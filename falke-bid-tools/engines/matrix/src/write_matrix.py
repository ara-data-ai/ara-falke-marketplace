"""
FALKE Matrix Pipeline — Excel Writer (Fresh Workbook)
======================================================
Writes normalized bid data into a brand-new openpyxl Workbook().

No template is loaded or copied — the FEB 26 file's merged cells, conditional
formatting, and drawing objects caused corruption on save.  This module
creates a clean xlsx that mirrors the FEB 26 row/column structure closely
enough for side-by-side value comparison.

Pipeline position:
    list[NormalizedBid]  →  write_matrix()  →  bid-comparison .xlsx

Layout:
  Col A  : CSI code / label
  Col B  : Row description
  Col C+ : Contractor groups, each 3 columns wide:
             +0  COST SUBTOTALS (main comparison number)
             +1  $/SF           (subtotal ÷ gsf)
             +2  blank separator

  Row 1  : Project title
  Row 2  : Project details
  Row 3  : blank
  Row 4  : Column headers (CSI / Building System / COST SUBTOTALS / $/SF / …)
  Row 5  : Contractor names
  Row 6  : Project label per contractor
  Row 7  : GSF per contractor
  Row 8  : blank
  Rows 9+: CSI division data
  …      : blank separator, footer section, blank separator, qualifications
"""

from __future__ import annotations

from decimal import Decimal
from math import ceil
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

from src.audit import AuditItem, AuditStatus
from src.normalized_models import (
    CellState,
    NormalizedBid,
    NormalizedDivision,
    ReclassRecommendation,
    grand_total_component_amounts,
)
from src.run_config import RunInputs

# Col C — the Normalization Note column on the Bid_Form mirror (Option C §2.1).
NOTE_COL = 3

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Project identity is a PER-RUN input (RunInputs), never hardcoded. These
# fallbacks exist only so a programmatic caller that passes a gsf but no
# RunInputs still produces a generic, non-client-specific title.
DEFAULT_SF_BASIS_LABEL = "GSF"

# Column widths
COL_A_WIDTH: float = 15.0
COL_B_WIDTH: float = 45.0
CONTRACTOR_COST_WIDTH: float = 18.0
CONTRACTOR_SF_WIDTH: float = 10.0
CONTRACTOR_SEP_WIDTH: float = 4.0

# Number format for dollar amounts
AMOUNT_FORMAT = "#,##0.00"

# ---------------------------------------------------------------------------
# Division and footer row definitions (FEB 26 CSI sequence)
# ---------------------------------------------------------------------------

DIVISION_ROWS: list[tuple[str, str]] = [
    ("DIV 01 00 00", "General Requirements"),
    ("DIV 02 00 00", "Existing Conditions"),
    ("DIV 03 00 00", "Concrete"),
    ("DIV 04 00 00", "Masonry"),
    ("DIV 05 00 00", "Metals"),
    ("DIV 06 00 00", "Wood, Plastics & Composites"),
    ("DIV 07 00 00", "Thermal & Moisture Protection"),
    ("DIV 08 00 00", "Openings"),
    ("DIV 09 00 00", "Finishes"),
    ("DIV 10 00 00", "Specialties"),
    ("DIV 11 00 00", "Equipment"),
    ("DIV 12 00 00", "Furnishings"),
    ("DIV 13 00 00", "Special Construction"),
    ("DIV 21 00 00", "Fire Suppression"),
    ("DIV 22 00 00", "Plumbing"),
    ("DIV 23 00 00", "HVAC"),
    ("DIV 25 00 00", "Integrated Automation"),
    ("DIV 26 00 00", "Electrical"),
    ("DIV 27 00 00", "Communications"),
    ("DIV 28 00 00", "Electronic Safety & Security"),
]

FOOTER_ROWS: list[tuple[str, str]] = [
    ("CONSTRUCTION_SUBTOTAL", "Construction Cost Subtotal"),
    ("GL_INSURANCE",          "General Liability Insurance"),
    ("BUILDERS_RISK",         "Builders Risk Insurance"),
    ("GC_FEE",                "GC Fee"),
    ("OVERHEAD_PROFIT",       "Overhead & Profit"),
    ("OTHER_FEES",            "Other Fees / Insurance"),
    ("FEES_SUBTOTAL",         "Fees Subtotal"),
    ("GRAND_TOTAL",           "GRAND TOTAL"),
    ("BOND",                  "Bond (Alternate)"),
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _to_float(amount: Optional[Decimal]) -> float:
    """Convert a Decimal (or None) to float. Returns 0.0 for None."""
    if amount is None:
        return 0.0
    return float(amount)


def _cell_amount(state: CellState, amount: Optional[Decimal]) -> float:
    """
    Return the numeric value to write for a CellValue, given its state.

    AMOUNT / EXPLICIT_ZERO / ALLOWANCE → float (or 0.0)
    NULL_BLANK / EXCLUDED / BY_OWNER_OTHERS → 0.0
    """
    if state in (CellState.AMOUNT, CellState.EXPLICIT_ZERO, CellState.ALLOWANCE):
        return _to_float(amount)
    return 0.0


def _sort_bids(bids: list[NormalizedBid]) -> list[NormalizedBid]:
    """
    Order columns by leveled_total ASCENDING — lowest leveled bid first, the
    natural board reading order, fully firm-agnostic (Marvin §2.5, Floyd C7).
    A bid with leveled_total None (e.g. no grand total extracted) sorts LAST so
    it never masquerades as the low bid. Ties break on contractor_name for
    determinism.
    """
    def _key(b: NormalizedBid) -> tuple[int, float, str]:
        lt = b.footer.leveled_total
        if lt is None:
            return (1, 0.0, b.contractor_name)
        return (0, float(lt), b.contractor_name)

    return sorted(bids, key=_key)


# Column A(1)=CSI, B(2)=Building System, C(3)=Normalization Note (Option C §2.1),
# then contractor groups start at D(4). Changing this ONE base offset shifts every
# contractor column right by one; reconcile.py reads _col_start back independently
# and inherits the shift (spec §7.2.5).
_CONTRACTOR_COL_BASE = 4  # D=4


def _col_start(contractor_index: int) -> int:
    """
    Return the 1-based openpyxl column index for the COST SUBTOTALS column
    of the given contractor (0-based index).

    Layout: A(1)=CSI, B(2)=description, C(3)=Normalization Note, D(4)=contractor 0.
    Each contractor group is 3 columns wide.
    """
    return _CONTRACTOR_COL_BASE + contractor_index * 3  # D=4, G=7, J=10 …


# ---------------------------------------------------------------------------
# Worksheet construction
# ---------------------------------------------------------------------------

def _write_header_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    bids: list[NormalizedBid],
    gsf: int,
    run: RunInputs,
) -> None:
    """Write rows 1–8: title, details, blank, column headers, contractor info.

    Project identity (title, details, per-contractor label) comes from the
    per-run RunInputs — never a hardcoded project (M1). The $/SF header carries
    the confirmed SF-basis label so the board knows what the denominator means
    (M2 / scoping §1.4).
    """
    bold = Font(bold=True)
    sf_label = run.sf_basis_label or DEFAULT_SF_BASIS_LABEL

    # Row 1 — project title
    ws.cell(row=1, column=1).value = f"{run.project_name} — Bid Comparison Matrix"
    ws.cell(row=1, column=1).font = bold

    # Row 2 — project details
    details = f"Project: {run.project_name} | {run.project_address} | {gsf:,.0f} {sf_label}"
    if run.rfp_label:
        details += f" | {run.rfp_label}"
    ws.cell(row=2, column=1).value = details

    # Row 3 — blank (intentional)

    # Row 4 — column headers
    ws.cell(row=4, column=1).value = "CSI"
    ws.cell(row=4, column=1).font = bold
    ws.cell(row=4, column=2).value = "Building System"
    ws.cell(row=4, column=2).font = bold
    # Col C — Normalization Note (Option C §2.1). Present on the Bid_Form mirror
    # so a known-firm reclass recommendation is on the FACE of the document.
    ws.cell(row=4, column=3).value = "Normalization Note"
    ws.cell(row=4, column=3).font = bold

    for i, bid in enumerate(bids):
        cost_col = _col_start(i)
        sf_col = cost_col + 1
        ws.cell(row=4, column=cost_col).value = "COST SUBTOTALS"
        ws.cell(row=4, column=cost_col).font = bold
        ws.cell(row=4, column=sf_col).value = f"$/{sf_label}"
        ws.cell(row=4, column=sf_col).font = bold

    # Row 5 — contractor names
    for i, bid in enumerate(bids):
        cost_col = _col_start(i)
        ws.cell(row=5, column=cost_col).value = bid.contractor_name
        ws.cell(row=5, column=cost_col).font = bold

    # Row 6 — project label per contractor
    for i in range(len(bids)):
        cost_col = _col_start(i)
        ws.cell(row=6, column=cost_col).value = run.project_name

    # Row 7 — SF basis per contractor
    for i in range(len(bids)):
        cost_col = _col_start(i)
        ws.cell(row=7, column=cost_col).value = gsf

    # Row 8 — blank (intentional)


def _descriptions_match(a: str, b: str) -> bool:
    """
    Return True when two line-item description strings are semantically the same.
    Uses word overlap ≥ 60% (no external libraries needed).
    """
    a_words = set(a.lower().split())
    b_words = set(b.lower().split())
    if not a_words or not b_words:
        return False
    overlap = len(a_words & b_words) / max(len(a_words), len(b_words))
    return overlap >= 0.6


def _lookup_item_amount(
    div: "NormalizedDivision | None",
    target_desc: str,
) -> Optional[float]:
    """
    Return the float amount for target_desc from a NormalizedDivision's
    line_item_cells.  Tries exact key first, then _descriptions_match.
    Returns None when no match or state maps to 0.0/blank.
    """
    if div is None:
        return None

    # Exact match first
    if target_desc in div.line_item_cells:
        cell = div.line_item_cells[target_desc]
        if cell.state in (CellState.AMOUNT, CellState.EXPLICIT_ZERO, CellState.ALLOWANCE):
            return _to_float(cell.amount)
        return None

    # Fuzzy match
    for key, cell in div.line_item_cells.items():
        if _descriptions_match(target_desc, key):
            if cell.state in (CellState.AMOUNT, CellState.EXPLICIT_ZERO, CellState.ALLOWANCE):
                return _to_float(cell.amount)
            return None

    return None


def _build_unified_descriptions(
    bids: list[NormalizedBid],
    csi_code: str,
) -> list[str]:
    """
    For a given CSI division, collect all unique line-item descriptions across
    all contractors in display order.

    Algorithm: iterate contractors in order; for each contractor's line_item_cells
    append a description only if no existing entry in the running list matches it
    (case-insensitive substring OR ≥60% word overlap).

    LUMP_SUM contractors with no real items (all NULL_BLANK) are skipped —
    their placeholder descriptions must not pollute the unified list when they
    carry no pricing signal.
    """
    all_descs: list[str] = []

    for bid in bids:
        div = _find_div(bid, csi_code)
        if div is None:
            continue

        # Skip divisions where all line items are NULL_BLANK (lump-sum placeholder rows)
        has_priced_item = any(
            cell.state in (CellState.AMOUNT, CellState.EXPLICIT_ZERO, CellState.ALLOWANCE)
            for cell in div.line_item_cells.values()
        )
        if not has_priced_item and div.cost_structure.value == "LUMP_SUM":
            continue

        for desc in div.line_item_cells:
            already_present = any(
                _descriptions_match(desc, existing) or desc.lower() in existing.lower()
                or existing.lower() in desc.lower()
                for existing in all_descs
            )
            if not already_present:
                all_descs.append(desc)

    return all_descs


def _find_div(bid: NormalizedBid, csi_code: str) -> "NormalizedDivision | None":
    """Return the first NormalizedDivision matching csi_code for a bid, or None."""
    for div in bid.divisions:
        if div.csi_code == csi_code:
            return div
    return None


def _marker_text(rec: ReclassRecommendation) -> str:
    """Build the in-place Normalization Note marker for a recommendation (§2.2).

    With a priced amount:
      ``Dumpsters $54,959 — normalize → DIV 01 (General Requirements).
        As-submitted here; applied in Leveled_Normalized.``
    Without an amount, the leading ``{desc} {amount} — `` is dropped.
    """
    to_short = _div_short(rec.to_division)
    tail = (
        f"normalize → {to_short} ({rec.to_division_name}). "
        f"As-submitted here; applied in Leveled_Normalized."
    )
    if rec.amount is not None:
        return f"{rec.line_item_desc} {_fmt_money(rec.amount)} — {tail}"
    return f"Normalize → {to_short} ({rec.to_division_name}). As-submitted here; applied in Leveled_Normalized."


def _div_short(csi_code: str) -> str:
    """Render the bare `DIV NN` form (drop ` 00 00`) for marker readability."""
    parts = csi_code.split()
    if len(parts) >= 2 and parts[0] == "DIV":
        return f"DIV {parts[1]}"
    return csi_code


def _fmt_money(amount: Decimal) -> str:
    """Board-display dollar string '$54,959' (matches normalize._fmt)."""
    return f"${int(amount.quantize(Decimal('1'))):,}"


def _note_by_desc(bids: list[NormalizedBid], csi_code: str) -> dict[str, ReclassRecommendation]:
    """Map a line-item description → its reclass recommendation for one division.

    A recommendation belongs to a description row on the mirror when its
    ``from_division`` equals the row's division (the dollars sit there as
    submitted). Keyed by the matched line_item_desc so the writer can stamp the
    Normalization Note on the right row. First match wins if two bidders share a
    description+target (the recommended target is identical by rule).
    """
    out: dict[str, ReclassRecommendation] = {}
    for bid in bids:
        for rec in bid.reclass_recommendations:
            if rec.from_division == csi_code and rec.line_item_desc not in out:
                out[rec.line_item_desc] = rec
    return out


def _write_division_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    bids: list[NormalizedBid],
    start_row: int,
    gsf: int,
    show_notes: bool = True,
) -> tuple[int, dict[str, int]]:
    """
    Write dynamic division rows starting at start_row.

    ``show_notes`` controls the Col C Normalization Note (Option C §2): True on
    the Bid_Form mirror (the recommendation is shown in place), False on the
    Leveled_Normalized sheet (the move is already applied there).

    For each CSI division:
      1. Division header row (bold): CSI code | Division name
      2. One row per unified line-item description (across all contractors)
      3. Bold subtotal row: blank | DIVISION NAME SUBTOTAL | sub amounts
      4. Blank spacer row

    Returns (next_row_after_divisions, subtotal_row_by_csi_code).
    subtotal_row_by_csi_code maps each CSI code to the Excel row number of its
    SUBTOTAL row — used by the caller to apply audit-driven color fills.
    """
    bold = Font(bold=True)

    # Build per-bid, per-division lookups:
    #   bid_div_lookup[bid_index][csi_code] = aggregated subtotal float
    bid_div_lookup: list[dict[str, float]] = []
    for bid in bids:
        aggregated: dict[str, float] = {}
        for div in bid.divisions:
            amount = _cell_amount(div.subtotal_cell.state, div.subtotal_cell.amount)
            aggregated[div.csi_code] = aggregated.get(div.csi_code, 0.0) + amount
        bid_div_lookup.append(aggregated)

    row = start_row
    subtotal_row_by_csi: dict[str, int] = {}

    for csi_code, div_name in DIVISION_ROWS:
        # --- Division header row (bold) ---
        c_csi = ws.cell(row=row, column=1)
        c_csi.value = csi_code
        c_csi.font = bold

        c_name = ws.cell(row=row, column=2)
        c_name.value = div_name
        c_name.font = bold

        row += 1

        # --- Gather per-bid NormalizedDivision objects for this CSI code ---
        bid_divs: list["NormalizedDivision | None"] = [
            _find_div(bid, csi_code) for bid in bids
        ]

        # --- Unified description list across all contractors ---
        all_descs = _build_unified_descriptions(bids, csi_code)

        # --- Normalization-Note recommendations for this division (Option C §2) ---
        notes_by_desc = _note_by_desc(bids, csi_code) if show_notes else {}

        # --- One row per line-item description ---
        for desc in all_descs:
            ws.cell(row=row, column=2).value = desc

            # Stamp the in-place Normalization Note (Col C) — text only, never
            # alters any subtotal; YELLOW recommend-review band on the note cell.
            rec = notes_by_desc.get(desc)
            if rec is not None:
                note_cell = ws.cell(row=row, column=NOTE_COL)
                note_cell.value = _marker_text(rec)
                note_cell.fill = YELLOW_FILL
                note_cell.alignment = Alignment(wrap_text=True)

            for i, div in enumerate(bid_divs):
                cost_col = _col_start(i)
                amount = _lookup_item_amount(div, desc)
                if amount is not None:
                    c = ws.cell(row=row, column=cost_col)
                    c.value = amount
                    c.number_format = AMOUNT_FORMAT

            row += 1

        # --- Subtotal row (bold) ---
        subtotal_label = div_name.upper() + " SUBTOTAL"
        c_sub_label = ws.cell(row=row, column=2)
        c_sub_label.value = subtotal_label
        c_sub_label.font = bold

        subtotal_row_by_csi[csi_code] = row  # record for audit-fill pass

        for i, bid in enumerate(bids):
            cost_col = _col_start(i)
            sf_col = cost_col + 1
            amount = bid_div_lookup[i].get(csi_code, 0.0)
            sf_val = round(amount / gsf, 2) if gsf > 0 else 0.0

            c_cost = ws.cell(row=row, column=cost_col)
            c_cost.value = amount
            c_cost.number_format = AMOUNT_FORMAT
            c_cost.font = bold

            c_sf = ws.cell(row=row, column=sf_col)
            c_sf.value = sf_val
            c_sf.number_format = AMOUNT_FORMAT
            c_sf.font = bold

        row += 1

        # --- Blank spacer ---
        row += 1

    return row, subtotal_row_by_csi  # next row after all divisions


def _write_footer_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    bids: list[NormalizedBid],
    start_row: int,
    gsf: int,
) -> tuple[int, list[dict]]:
    """
    Write a blank separator then the footer rows (construction subtotal,
    insurance, GC fee, overhead & profit, other fees, fees subtotal, grand
    total, bond).

    Every component row that COMPOSES the grand total — GL, Builders Risk, GC
    fee, Overhead & Profit, and Other Fees/Insurance — is rendered as a labeled
    row, so the footer visibly ties to the grand total for a board. The amounts
    come from grand_total_component_amounts() (the single source of truth shared
    with audit.py and reconcile.py): a contractor that folds insurance into
    `other_fees_subtotal` (e.g. PBS) shows it on the Other Fees row, while a memo
    `other_fees` line that merely duplicates fees already counted is rendered as
    0 so it is not double-counted. FEES_SUBTOTAL is the sum of every additive
    fee component, so CONSTRUCTION_SUBTOTAL + FEES_SUBTOTAL == GRAND_TOTAL.

    Returns (next_row_after_footer, list[per_bid_footer_summary]).
    """
    bold = Font(bold=True)
    row = start_row + 1  # +1 blank separator

    # Pre-resolve the additive grand-total composition per bid (single source of
    # truth) so the rendered component rows and FEES_SUBTOTAL agree with audit /
    # reconcile, and a memo other_fees is suppressed to 0.
    components = [grand_total_component_amounts(bid.footer) for bid in bids]

    # The additive fee components that roll up into FEES_SUBTOTAL (everything
    # composing the grand total EXCEPT construction). Bond is an alternate, not
    # part of the base grand total, so it is excluded here.
    _FEE_COMPONENT_KEYS = (
        "GL_INSURANCE",
        "BUILDERS_RISK",
        "GC_FEE",
        "OVERHEAD_PROFIT",
        "OTHER_FEES",
    )

    summaries: list[dict] = []
    for bid in bids:
        summaries.append({})

    for key, label in FOOTER_ROWS:
        ws.cell(row=row, column=1).value = key
        ws.cell(row=row, column=2).value = label
        if key == "GRAND_TOTAL":
            ws.cell(row=row, column=2).font = bold

        for i, bid in enumerate(bids):
            cost_col = _col_start(i)
            footer = bid.footer
            comp = components[i]

            if key == "CONSTRUCTION_SUBTOTAL":
                val = float(comp.get("CONSTRUCTION_SUBTOTAL", 0))
            elif key in _FEE_COMPONENT_KEYS:
                # Additive contribution per the shared composition (memo
                # other_fees → absent from comp → 0).
                val = float(comp.get(key, 0))
            elif key == "FEES_SUBTOTAL":
                val = sum(
                    summaries[i].get(k, 0.0) for k in _FEE_COMPONENT_KEYS
                )
            elif key == "GRAND_TOTAL":
                val = _cell_amount(footer.grand_total.state, footer.grand_total.amount)
            elif key == "BOND":
                val = _cell_amount(footer.bond.state, footer.bond.amount)
            else:
                val = 0.0

            summaries[i][key] = val

            c = ws.cell(row=row, column=cost_col)
            c.value = val
            c.number_format = AMOUNT_FORMAT
            if key == "GRAND_TOTAL":
                c.font = bold

        row += 1

    return row, summaries


def _write_alternates(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    bids: list[NormalizedBid],
    start_row: int,
) -> int:
    """Write bid alternates in their OWN clearly-labeled section (M7).

    Alternates (add/deduct options) are NEVER folded into the base/leveled
    total — the base comparison stays apples-to-apples. Each contractor's
    alternates are listed under their column. If no bidder submitted any
    alternate, the section is omitted entirely. Returns the next free row.
    """
    if not any(bid.footer.alternates for bid in bids):
        return start_row

    bold = Font(bold=True)
    row = start_row + 1  # blank separator

    ws.cell(row=row, column=1).value = "ALTERNATES"
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=2).value = (
        "Bid Alternates (add/deduct — NOT included in base comparison)"
    )
    ws.cell(row=row, column=2).font = bold
    row += 1

    # One row per (contractor, alternate). Description in col B, amount under the
    # contractor's cost column — kept visually separate from the base divisions.
    for i, bid in enumerate(bids):
        cost_col = _col_start(i)
        for alt in bid.footer.alternates:
            ws.cell(row=row, column=1).value = bid.contractor_name
            ws.cell(row=row, column=2).value = alt.description
            c = ws.cell(row=row, column=cost_col)
            if alt.amount is not None:
                c.value = float(alt.amount)
                c.number_format = AMOUNT_FORMAT
            else:
                c.value = alt.display
            row += 1

    return row


def _write_qualifications(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    bids: list[NormalizedBid],
    start_row: int,
) -> None:
    """Write one qualifications row per contractor, separated by a blank row."""
    row = start_row + 1  # +1 blank separator

    ws.cell(row=row, column=1).value = "QUALIFICATIONS"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = "Contractor Qualifications"
    ws.cell(row=row, column=2).font = Font(bold=True)

    row += 1
    for i, bid in enumerate(bids):
        cost_col = _col_start(i)
        ws.cell(row=row, column=cost_col - 1).value = bid.contractor_name
        qual_cell = ws.cell(row=row, column=cost_col)
        qual_cell.value = bid.qualifications_text or ""
        qual_cell.alignment = Alignment(wrap_text=True)


def _set_column_widths(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    num_contractors: int,
) -> None:
    """Set column widths for readability."""
    from openpyxl.utils import get_column_letter

    ws.column_dimensions["A"].width = COL_A_WIDTH
    ws.column_dimensions["B"].width = COL_B_WIDTH

    for i in range(num_contractors):
        cost_col = _col_start(i)
        sf_col   = cost_col + 1
        sep_col  = cost_col + 2

        ws.column_dimensions[get_column_letter(cost_col)].width = CONTRACTOR_COST_WIDTH
        ws.column_dimensions[get_column_letter(sf_col)].width   = CONTRACTOR_SF_WIDTH
        ws.column_dimensions[get_column_letter(sep_col)].width  = CONTRACTOR_SEP_WIDTH


# ---------------------------------------------------------------------------
# Audit sheet fills and AUDIT worksheet
# ---------------------------------------------------------------------------

# Cell fill constants — PatternFill is safe on fresh workbooks (no existing styles to corrupt)
RED_FILL    = PatternFill("solid", fgColor="FFCCCC")   # soft red
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")   # soft yellow
GREEN_FILL  = PatternFill("solid", fgColor="CCFFCC")   # soft green

_STATUS_FILL = {
    AuditStatus.RED:    RED_FILL,
    AuditStatus.YELLOW: YELLOW_FILL,
    AuditStatus.GREEN:  GREEN_FILL,
}

_STATUS_SORT_KEY = {
    AuditStatus.RED:    0,
    AuditStatus.YELLOW: 1,
    AuditStatus.GREEN:  2,
}


def _apply_audit_fills(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    bids: list[NormalizedBid],
    subtotal_row_by_csi: dict[str, int],
    audit_items: list[AuditItem],
    view: str = "mirror",
) -> None:
    """
    Color-code SUBTOTAL cells in a data sheet based on audit findings.

    A division subtotal cell is colored by the WORST status of any division-
    scoped AuditItem for that (contractor, division): RED wins over YELLOW;
    GREEN triggers no fill. ``view`` selects this sheet's slice (Option C §4):
    only items whose ``view`` is ``both`` or equals this sheet's ``view`` apply.

    On the mirror (``view="mirror"``) the KNOWN_FIRM_RECLASSIFIED YELLOW belongs
    on the Normalization Note cell (stamped during division writing), NOT on the
    division subtotal — the as-submitted subtotal is correct and must not be
    flagged as if something is wrong with it (§2.4). So that code is excluded
    from the subtotal-fill index here.
    """
    # Build index: (contractor_name, csi_code) → worst AuditStatus.
    from openpyxl.utils import get_column_letter
    from src.audit import AuditCode

    _STATUS_RANK = {AuditStatus.RED: 0, AuditStatus.YELLOW: 1, AuditStatus.GREEN: 2}
    worst: dict[tuple[str, str], AuditStatus] = {}
    for item in audit_items:
        if item.division_csi is None:
            continue
        if item.view not in ("both", view):
            continue
        if view == "mirror" and item.code == AuditCode.KNOWN_FIRM_RECLASSIFIED:
            # YELLOW lives on the Normalization Note cell on the mirror, not here.
            continue
        key = (item.contractor_name, item.division_csi)
        if key not in worst or _STATUS_RANK[item.status] < _STATUS_RANK[worst[key]]:
            worst[key] = item.status

    for i, bid in enumerate(bids):
        cost_col = _col_start(i)
        col_letter = get_column_letter(cost_col)
        name = bid.contractor_name

        for csi_code, subtotal_row in subtotal_row_by_csi.items():
            status = worst.get((name, csi_code))
            if status == AuditStatus.RED:
                ws[f"{col_letter}{subtotal_row}"].fill = RED_FILL
            elif status == AuditStatus.YELLOW:
                ws[f"{col_letter}{subtotal_row}"].fill = YELLOW_FILL
            # GREEN / None: no fill change (default)


def _write_audit_sheet(
    wb: openpyxl.Workbook,
    audit_items: list[AuditItem],
) -> None:
    """
    Create and populate the AUDIT worksheet in wb.

    Layout:
      Row 1: Title
      Row 2: Subtitle
      Row 3: blank
      Row 4: Column headers
      Row 5+: One row per AuditItem sorted RED→YELLOW→GREEN, then contractor, then division
      (2 blank rows after last item)
      Summary block: totals by status
    """
    ws = wb.create_sheet(title="AUDIT")
    bold = Font(bold=True)

    # --- Column widths ---
    from openpyxl.utils import get_column_letter
    col_widths = [10, 14, 28, 30, 16, 35, 18, 60]  # A through H (View col inserted)
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Row 1: Title ---
    ws.cell(row=1, column=1).value = (
        "FALKE Matrix — Extraction & Normalization Audit Report"
    )
    ws.cell(row=1, column=1).font = bold

    # --- Row 2: Subtitle ---
    ws.cell(row=2, column=1).value = (
        "Generated by ARA Pipeline | Items requiring action before bid award"
    )

    # --- Row 3: blank ---

    # --- Row 4: Column headers (View column inserted between Status and Code, §4) ---
    headers = ["Status", "View", "Code", "Contractor", "Division", "Line Item", "Value", "Message"]
    for col_idx, header in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx)
        c.value = header
        c.font = bold

    # --- Sort items: RED first, then YELLOW, then GREEN; then by contractor; then division ---
    sorted_items = sorted(
        audit_items,
        key=lambda a: (
            _STATUS_SORT_KEY[a.status],
            a.contractor_name,
            a.division_csi or "",
        ),
    )

    _VIEW_LABEL = {"leveled": "Leveled", "mirror": "As-Submitted", "both": "Both"}

    # --- Rows 5+: One row per AuditItem ---
    data_start_row = 5
    for row_offset, item in enumerate(sorted_items):
        row = data_start_row + row_offset
        fill = _STATUS_FILL[item.status]

        values = [
            item.status.value,
            _VIEW_LABEL.get(item.view, item.view),
            item.code.value,
            item.contractor_name,
            item.division_csi or "",
            item.line_item_desc or "",
            item.value or "",
            item.message,
        ]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col_idx)
            c.value = val
            c.fill = fill
            # Bold the Status cell text
            if col_idx == 1:
                c.font = Font(bold=True)

    # --- Summary block (2 blank rows after last data row) ---
    last_data_row = data_start_row + len(sorted_items) - 1
    summary_start = last_data_row + 3  # 2 blank rows then summary

    red_count    = sum(1 for a in audit_items if a.status == AuditStatus.RED)
    yellow_count = sum(1 for a in audit_items if a.status == AuditStatus.YELLOW)
    green_count  = sum(1 for a in audit_items if a.status == AuditStatus.GREEN)
    total_count  = len(audit_items)

    summary_lines = [
        (f"Total items audited:   {total_count}", None),
        (f"RED Critical:          {red_count}  — must resolve before award", RED_FILL),
        (f"YELLOW Review:         {yellow_count}  — verify before finalizing", YELLOW_FILL),
        (f"GREEN Verified:        {green_count}  — clean", GREEN_FILL),
    ]
    for i, (text, fill) in enumerate(summary_lines):
        c = ws.cell(row=summary_start + i, column=1)
        c.value = text
        c.font = bold
        if fill:
            c.fill = fill


# ---------------------------------------------------------------------------
# Leveled-view banner (Option C §3.3)
# ---------------------------------------------------------------------------

# Two banner lines, written into rows 1–2 of Leveled_Normalized (replacing the
# normal title/details). Rows 4–8 (headers, names, GSF) stay at the SAME rows as
# the mirror so reconcile.py's row-5 name read and label-anchored reads work on
# both sheets unchanged.
_LEVELED_BANNER_LINE_1 = (
    "ESTIMATOR-NORMALIZED VIEW — does NOT match the submitted bids. Dollars have "
    "been moved between divisions for apples-to-apples comparison. See the "
    "Bid_Form sheet for each bid exactly as submitted."
)
_LEVELED_BANNER_LINE_2 = (
    "Normalization applied: known-firm division reclassifications (see "
    "Normalization Note column on Bid_Form for each move and its rationale)."
)


def _populate_data_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    ordered_bids: list[NormalizedBid],
    gsf: int,
    run: RunInputs,
    audit_items: Optional[list[AuditItem]],
    view: str,
    show_notes: bool,
) -> list[dict]:
    """Fill one data worksheet (mirror or leveled) end-to-end.

    Returns the per-bid footer summaries (used only for the mirror's report).
    ``view`` is ``"mirror"`` or ``"leveled"`` and selects the audit-fill slice;
    ``show_notes`` writes the Col C Normalization Note (mirror only).
    """
    _write_header_rows(ws, ordered_bids, gsf, run)

    if len(ordered_bids) == 1:
        notice = ws.cell(row=3, column=1)
        notice.value = "Single bid — no competitive comparison available."
        notice.font = Font(bold=True, italic=True)

    DIVISION_START_ROW = 9
    next_row, subtotal_row_by_csi = _write_division_rows(
        ws, ordered_bids, DIVISION_START_ROW, gsf, show_notes=show_notes
    )

    if audit_items:
        _apply_audit_fills(ws, ordered_bids, subtotal_row_by_csi, audit_items, view=view)

    next_row, footer_summaries = _write_footer_rows(ws, ordered_bids, next_row, gsf)
    next_row = _write_alternates(ws, ordered_bids, next_row)
    _write_qualifications(ws, ordered_bids, next_row)
    _set_column_widths(ws, len(ordered_bids))
    return footer_summaries


def _write_leveled_banner(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    num_contractors: int,
) -> None:
    """Write the unmissable two-line leveled banner into rows 1–2 (§3.3).

    Bold line 1 + provenance line 2, both YELLOW so they read as "interpreted,
    verify". Written AFTER the header so it overrides the title/details cells.
    """
    used_width = _col_start(max(num_contractors - 1, 0)) + 2
    for row, (text, weight_bold) in enumerate(
        [(_LEVELED_BANNER_LINE_1, True), (_LEVELED_BANNER_LINE_2, False)], start=1
    ):
        c = ws.cell(row=row, column=1)
        c.value = text
        c.font = Font(bold=weight_bold)
        c.alignment = Alignment(wrap_text=True)
        # YELLOW across the used width so the banner is visible end-to-end.
        for col in range(1, used_width + 1):
            ws.cell(row=row, column=col).fill = YELLOW_FILL


# ---------------------------------------------------------------------------
# Stage 6b LOUD QUARANTINE — post-reconcile annotation pass
# ---------------------------------------------------------------------------
#
# When Stage 6b (reconcile_written_matrix) returns ≥1 POST_WRITE_TIEOUT_FAILURE,
# the matrix is STILL delivered (Derick's decision) but every affected figure is
# loud-quarantined so a non-technical board cannot mistake it for a verified
# matrix (Marvin's STAGE6B-QUARANTINE-DISCLOSURE-SPEC.md). The disclosure is three
# stacked, board-facing RED signals:
#   1. a RED banner at the top of Bid_Form AND Leveled_Normalized (this module),
#   2. a RED fill + verify-against-source comment on each failing cell, and
#   3. a RED AUDIT row + a QUARANTINE summary line (rows added by the pipeline via
#      the audit_items list; the summary line is written here).
#
# This runs AFTER reconcile so the banner reflects the tie-out result. The banner
# is written into rows 1–3 by OVERRIDING the existing title/details cells (same
# override idiom as _write_leveled_banner) — it does NOT shift the data rows, so
# the contractor-name row (5), footer label rows, and SUBTOTAL label rows that
# reconcile.py read stay at their original positions and the cell marks below
# re-locate them by the same labels (Marvin §2: "rows reconcile.py reads stay
# anchored by label, not by absolute number"). On Leveled_Normalized the RED
# quarantine banner sits ABOVE the existing yellow normalization banner.

# Banner text — Marvin §2 / §6 (exact strings). {N} filled per workbook.
_QUARANTINE_BANNER_LINE_1 = (
    "⚠ AUTOMATED CHECK FAILED — DO NOT RELY ON THE FLAGGED FIGURES FOR AN AWARD "
    "DECISION."
)


def _quarantine_banner_line_2(n: Optional[int]) -> str:
    """Marvin §2 Line 2 with singular/plural and the structural fallback.

    ``n`` is the count of distinct flagged figures; ``None`` ⇒ the count cannot be
    cleanly enumerated (a structural failure), so the banner uses "one or more
    figures" and never under-counts itself into looking minor.
    """
    if n is None:
        figures = "one or more figures"
        verb = "do"
    else:
        figures = f"{n} figure{'' if n == 1 else 's'}"
        verb = "does" if n == 1 else "do"
    return (
        f"This bid-comparison matrix did not pass the tool's final self-check. "
        f"{figures} on this sheet {verb} not reconcile to the source bids — the "
        f"tool's written value does not match its own verified calculation. This "
        f"is a tool/formatting problem, not a finding about any contractor's bid."
    )


_QUARANTINE_BANNER_LINE_3 = (
    "Before this matrix is used to award, a person must verify each flagged figure "
    "directly against that contractor's submitted bid. Flagged figures are marked "
    'in their cells with "⚠ does not reconcile to source — verify". The full list '
    "is on the AUDIT tab (filter the Code column for POST_WRITE_TIEOUT_FAILURE). "
    "Do not award off this matrix until every flagged figure has been checked by "
    "hand."
)

# Cell-comment text — Marvin §2 / §6 (exact string).
_QUARANTINE_CELL_COMMENT = (
    "⚠ does not reconcile to source — verify. The tool wrote {written} here; its "
    "own verified calculation was {expected} (difference {delta}). Check this "
    "figure against {contractor}'s submitted bid before relying on it."
)

# AUDIT summary quarantine line — Marvin §4 / §6 (exact string).
_QUARANTINE_AUDIT_SUMMARY = (
    "QUARANTINE: {n} figure(s) failed the tool's self-check and are FLAGGED on the "
    "Bid_Form sheet. Verify each against the contractor's submitted bid before any "
    "award."
)

# Message substrings that map a POST_WRITE_TIEOUT_FAILURE to a markable cell.
_GT_CELL_MARKERS = (
    "Grand-total tie-out FAILED",
    "Footer arithmetic FAILED",
    "Mirror/leveled grand-total mismatch",
)
_DIV_CELL_MARKER = "Division subtotal tie-out FAILED"

# The footer-component rows that compose the grand total (same set reconcile.py
# re-sums for check 2). Defined locally to avoid a circular import (reconcile
# imports from this module). Bond is an alternate, not part of the base total.
_GRAND_TOTAL_COMPONENT_KEYS = (
    "CONSTRUCTION_SUBTOTAL",
    "GL_INSURANCE",
    "BUILDERS_RISK",
    "GC_FEE",
    "OVERHEAD_PROFIT",
    "OTHER_FEES",
)


def _used_width_chars(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    used_width: int,
) -> float:
    """Approximate the character capacity of one wrapped line across the merged
    banner (sum of the set column widths from A through ``used_width``).

    openpyxl column width is roughly in characters of the default font, so the
    sum is a usable estimate of how many characters fit on one line of the
    full-width merged banner — used to size the row height so the text shows
    horizontally instead of stacking in column A.
    """
    from openpyxl.utils import get_column_letter

    total = 0.0
    for col in range(1, used_width + 1):
        dim = ws.column_dimensions.get(get_column_letter(col))
        total += (dim.width if dim is not None and dim.width else 8.43)
    return total


def _write_quarantine_banner(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    num_contractors: int,
    n: Optional[int],
    start_row: int = 1,
) -> None:
    """Write the RED 3-row quarantine banner into ``start_row .. start_row+2``.

    Each banner row is MERGED across the full used width (col A → last contractor
    column) so the long L2/L3 paragraphs flow HORIZONTALLY as a proper full-width
    banner instead of stacking into a tall, narrow column-A block. Each merged row
    gets ``wrap_text`` + an explicit row height sized to the text so it is fully
    visible. Solid RED across the used width, Line 1 bold. Overrides whatever is in
    those rows (title/details on Bid_Form; the writer leaves rows 1–3 for it on
    Leveled_Normalized). Does not shift data rows.
    """
    from openpyxl.utils import get_column_letter

    used_width = _col_start(max(num_contractors - 1, 0)) + 2
    last_col = get_column_letter(used_width)
    line_chars = _used_width_chars(ws, used_width)
    lines = [
        (_QUARANTINE_BANNER_LINE_1, True),
        (_quarantine_banner_line_2(n), False),
        (_QUARANTINE_BANNER_LINE_3, False),
    ]
    for offset, (text, weight_bold) in enumerate(lines):
        row = start_row + offset
        c = ws.cell(row=row, column=1)
        c.value = text
        c.font = Font(bold=weight_bold)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        # RED across the used width FIRST so every underlying cell is filled even
        # before the merge collapses them visually.
        for col in range(1, used_width + 1):
            ws.cell(row=row, column=col).fill = RED_FILL
        # MERGE the row across the full used width so the text flows horizontally.
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=used_width)
        # Explicit row height: estimate wrapped-line count from the text length vs
        # the merged width's character capacity, ~15 pts per line (default font).
        est_lines = max(1, ceil(len(text) / max(line_chars, 1.0)))
        ws.row_dimensions[row].height = est_lines * 15.0


def _mark_cell(ws, row: int, col: int, written: str, expected: str,
               delta: str, contractor: str) -> None:
    """RED fill + verify-against-source comment on a single failing cell (§2)."""
    cell = ws.cell(row=row, column=col)
    cell.fill = RED_FILL
    cell.comment = Comment(
        _QUARANTINE_CELL_COMMENT.format(
            written=written, expected=expected, delta=delta, contractor=contractor,
        ),
        "FALKE Stage 6b",
    )


def _fmt_q(amount: Decimal) -> str:
    """Board-display dollar string for the cell comment (matches reconcile._fmt)."""
    return f"${int(amount.quantize(Decimal('1'))):,}"


def _mark_failing_cells(
    ws,
    sheet_name: str,
    ordered_bids: list[NormalizedBid],
    failures: list[AuditItem],
) -> None:
    """Mark each failing GRAND_TOTAL / SUBTOTAL cell on ``ws`` (Marvin §2/§3).

    Cells are re-located by label (contractor name row 5, GRAND_TOTAL col-A label,
    ``… SUBTOTAL`` col-B labels) — the same anchors reconcile.py uses — and the
    written/expected/delta are recomputed here from the workbook + the blessed
    bids (writer-independent), so a parse of reconcile's message string is never
    required. Structural failures (no single wrong cell) are skipped here; they are
    disclosed by the banner count and the AUDIT row only.
    """
    from openpyxl.utils import get_column_letter  # noqa: F401  (kept local, parity)

    name_to_col = _find_contractor_cols(ws, ordered_bids)
    grand_total_row = _find_label_row_col_a(ws, "GRAND_TOTAL")
    subtotal_rows = _find_subtotal_label_rows(ws)
    component_rows = {
        key: _find_label_row_col_a(ws, key)
        for key in _GRAND_TOTAL_COMPONENT_KEYS
    }
    blessed_by_name = {b.contractor_name: b for b in ordered_bids}

    # A GRAND_TOTAL cell can be implicated by more than one failure (e.g. both a
    # grand-total tie-out and a mirror mismatch). Mark each GT cell ONCE, choosing
    # the "expected" that best describes the defect: grand-total tie-out (blessed
    # GT) > mirror mismatch (blessed GT) > footer arithmetic (the summed
    # components). Collect per-contractor, then emit.
    gt_priority = {"GRAND": 0, "MISMATCH": 1, "FOOTER": 2}
    gt_marks: dict[str, tuple[int, Decimal]] = {}  # contractor → (priority, expected)

    for f in failures:
        # Only failures stamped for THIS sheet (or the cross-sheet GT mismatch,
        # which has no [sheet] prefix and applies to GRAND_TOTAL on both sheets).
        is_mismatch = "Mirror/leveled grand-total mismatch" in f.message
        if not is_mismatch and f"[{sheet_name}]" not in f.message:
            continue

        col = name_to_col.get(f.contractor_name)
        if col is None:
            continue
        bid = blessed_by_name.get(f.contractor_name)
        if bid is None:
            continue

        # Division-subtotal failure → mark the (contractor, division) SUBTOTAL cell.
        if f.division_csi and _DIV_CELL_MARKER in f.message:
            sub_row = subtotal_rows.get(f.division_csi)
            if sub_row is None:
                continue
            written = _as_dec(ws.cell(row=sub_row, column=col).value)
            expected = _blessed_div_subtotal(bid, f.division_csi)
            _mark_cell(
                ws, sub_row, col,
                _fmt_q(written), _fmt_q(expected), _fmt_q(abs(written - expected)),
                f.contractor_name,
            )
            continue

        # Grand-total / footer-arithmetic / mirror-mismatch → GRAND_TOTAL cell.
        if grand_total_row is None:
            continue
        if "Grand-total tie-out FAILED" in f.message:
            kind, expected = "GRAND", _blessed_grand_total(bid)
        elif is_mismatch:
            kind, expected = "MISMATCH", _blessed_grand_total(bid)
        elif "Footer arithmetic FAILED" in f.message:
            # "Verified calculation" here is the sum of the written components —
            # that is what the GRAND TOTAL should equal but doesn't.
            kind = "FOOTER"
            expected = sum(
                (
                    _as_dec(ws.cell(row=component_rows[key], column=col).value)
                    for key in _GRAND_TOTAL_COMPONENT_KEYS
                    if component_rows.get(key) is not None
                ),
                Decimal("0"),
            )
        else:
            continue  # structural — no single cell (banner + AUDIT cover it)

        prev = gt_marks.get(f.contractor_name)
        if prev is None or gt_priority[kind] < prev[0]:
            gt_marks[f.contractor_name] = (gt_priority[kind], expected)

    # Emit one mark per implicated GRAND_TOTAL cell.
    if grand_total_row is not None:
        for name, (_prio, expected) in gt_marks.items():
            col = name_to_col[name]
            written = _as_dec(ws.cell(row=grand_total_row, column=col).value)
            _mark_cell(
                ws, grand_total_row, col,
                _fmt_q(written), _fmt_q(expected), _fmt_q(abs(written - expected)),
                name,
            )


def _find_name_row(ws) -> int:
    """Return the contractor-name row: the row just below the "CSI" header row.

    Anchored to the "CSI" label in col A (header row) so it survives a banner
    row-shift on Leveled_Normalized; falls back to row 5 (the writer's default).
    """
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "CSI":
            return row + 1
    return 5


def _find_contractor_cols(ws, bids: list[NormalizedBid]) -> dict[str, int]:
    """Map contractor name → its COST column (read back from the name row)."""
    name_row = _find_name_row(ws)
    out: dict[str, int] = {}
    for i in range(len(bids)):
        col = _col_start(i)
        name = ws.cell(row=name_row, column=col).value
        if isinstance(name, str) and name:
            out[name] = col
    return out


def _find_label_row_col_a(ws, key: str) -> Optional[int]:
    """Row whose col-A value equals a footer key (e.g. GRAND_TOTAL)."""
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == key:
            return row
    return None


def _find_subtotal_label_rows(ws) -> dict[str, int]:
    """Map each division CSI code → its SUBTOTAL row (read back from col B)."""
    label_to_code = {
        f"{name.upper()} SUBTOTAL": code for code, name in DIVISION_ROWS
    }
    out: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row=row, column=2).value
        if isinstance(label, str) and label in label_to_code:
            out[label_to_code[label]] = row
    return out


def _as_dec(v: object) -> Decimal:
    """Coerce a cell value to Decimal; non-numeric → 0 (matches reconcile._as_decimal)."""
    if v is None:
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    try:
        return Decimal(str(v))
    except Exception:
        return Decimal("0")


def _blessed_grand_total(bid: NormalizedBid) -> Decimal:
    gt = bid.footer.grand_total
    if gt.amount is not None and gt.state == CellState.AMOUNT:
        return Decimal(str(gt.amount))
    return Decimal("0")


def _blessed_div_subtotal(bid: NormalizedBid, csi_code: str) -> Decimal:
    total = Decimal("0")
    for div in bid.divisions:
        if div.csi_code != csi_code:
            continue
        cell = div.subtotal_cell
        if cell.state in (CellState.AMOUNT, CellState.EXPLICIT_ZERO,
                          CellState.ALLOWANCE) and cell.amount is not None:
            total += Decimal(str(cell.amount))
    return total


def _quarantine_figure_count(failures: list[AuditItem]) -> Optional[int]:
    """Distinct flagged-figure count for the banner ``{N}`` (Marvin §2).

    Counts one per failing GRAND_TOTAL (grand-total / footer-arithmetic /
    mirror-mismatch) and one per failing (contractor, division) SUBTOTAL. Returns
    ``None`` when a STRUCTURAL failure is present (missing row/sheet/column or
    audit-row parity) — the count can't be cleanly enumerated, so the banner falls
    back to "one or more figures" and never under-counts.
    """
    figures: set[tuple[str, str]] = set()
    structural = False
    for f in failures:
        if f.division_csi and _DIV_CELL_MARKER in f.message:
            figures.add((f.contractor_name, f.division_csi))
        elif any(m in f.message for m in _GT_CELL_MARKERS):
            figures.add((f.contractor_name, "GRAND_TOTAL"))
        else:
            structural = True
    if structural:
        return None
    return len(figures)


_AUDIT_VIEW_LABEL = {"leveled": "Leveled", "mirror": "As-Submitted", "both": "Both"}


def _append_audit_failure_rows(
    wb: openpyxl.Workbook,
    failures: list[AuditItem],
) -> None:
    """Append the POST_WRITE_TIEOUT_FAILURE rows to the AUDIT sheet (Marvin §4).

    The banner Line 3 tells the board to "filter the Code column for
    POST_WRITE_TIEOUT_FAILURE" — so each tie-out failure must appear as a RED row
    on the AUDIT tab. Rows are inserted at the END of the existing data region
    (just before the 2 blank rows + summary block), matching the 8-column layout
    written by ``_write_audit_sheet`` (Status, View, Code, Contractor, Division,
    Line Item, Value, Message). All tie-out failures are RED.

    NOTE: this runs AFTER reconcile's check-4 (audit-row parity), which counted
    only the Stage-5b rows — so appending here does not retroactively trip parity.
    """
    if "AUDIT" not in wb.sheetnames or not failures:
        return
    ws = wb["AUDIT"]

    # Locate the end of the contiguous data region (col-A non-empty from row 5).
    row = 5
    while ws.cell(row=row, column=1).value not in (None, ""):
        row += 1
    insert_at = row  # first blank row after the data region

    ws.insert_rows(insert_at, len(failures))
    for offset, f in enumerate(failures):
        r = insert_at + offset
        values = [
            f.status.value,
            _AUDIT_VIEW_LABEL.get(f.view, f.view),
            f.code.value,
            f.contractor_name,
            f.division_csi or "",
            f.line_item_desc or "",
            f.value or "",
            f.message,
        ]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=r, column=col_idx)
            c.value = val
            c.fill = RED_FILL
            if col_idx == 1:
                c.font = Font(bold=True)


def _append_audit_quarantine_line(wb: openpyxl.Workbook, n: Optional[int]) -> None:
    """Add the QUARANTINE summary line ABOVE the RED/YELLOW/GREEN tally (§4).

    Inserts a RED-filled line at the top of the existing summary block on the
    AUDIT sheet. ``n`` ⇒ the distinct flagged-figure count; ``None`` ⇒ "one or
    more".
    """
    if "AUDIT" not in wb.sheetnames:
        return
    ws = wb["AUDIT"]
    # Find the summary block: the first "Total items audited:" row (col A).
    summary_row = None
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        if isinstance(v, str) and v.startswith("Total items audited:"):
            summary_row = row
            break
    count_str = "one or more" if n is None else str(n)
    text = _QUARANTINE_AUDIT_SUMMARY.format(n=count_str)
    if summary_row is None:
        # No summary block (no audit_items) — append at the next free row.
        summary_row = ws.max_row + 2
        c = ws.cell(row=summary_row, column=1)
    else:
        ws.insert_rows(summary_row, 1)
        c = ws.cell(row=summary_row, column=1)
    c.value = text
    c.font = Font(bold=True)
    c.fill = RED_FILL


def apply_quarantine(
    output_path: str | Path,
    failures: list[AuditItem],
    bids: list[NormalizedBid],
    leveled_bids: Optional[list[NormalizedBid]] = None,
) -> int:
    """Loud-quarantine the just-written workbook IN PLACE (Marvin's spec).

    Re-opens ``output_path``, writes the RED banner on Bid_Form + Leveled_Normalized,
    marks each failing cell (RED fill + verify comment), appends the RED
    POST_WRITE_TIEOUT_FAILURE rows + the QUARANTINE summary line to the AUDIT
    sheet, and re-saves. Call ONLY when ``failures`` is non-empty. Returns
    the distinct flagged-figure count actually rendered in the banner ``{N}``
    (``-1`` when the structural fallback "one or more figures" was used) for the
    pipeline's console summary.

    Cell-mark re-location is label-anchored (writer-independent), matching
    reconcile.py — so it survives the banner write (which does not shift rows).
    """
    output_path = Path(output_path)
    wb = openpyxl.load_workbook(output_path)

    ordered_mirror = _sort_bids(bids)
    lev_by_name = {b.contractor_name: b for b in (leveled_bids or bids)}
    ordered_leveled = [
        lev_by_name.get(b.contractor_name, b) for b in ordered_mirror
    ]

    n = _quarantine_figure_count(failures)
    num_contractors = len(ordered_mirror)

    # --- Banner + cell marks on Bid_Form ---
    if "Bid_Form" in wb.sheetnames:
        ws = wb["Bid_Form"]
        _write_quarantine_banner(ws, num_contractors, n, start_row=1)
        _mark_failing_cells(ws, "Bid_Form", ordered_mirror, failures)

    # --- Banner + cell marks on Leveled_Normalized (RED banner ABOVE yellow) ---
    # The leveled sheet already carries the yellow normalization banner in rows
    # 1–2. Insert 3 rows at the top so the RED quarantine banner sits ABOVE it
    # (Marvin §2: "the RED quarantine block sits above the yellow one"). This
    # shifts the leveled sheet's data down by 3 — harmless, because reconcile.py
    # already ran on the pre-quarantine file, and the cell marks below re-locate
    # cells by label (name row anchored to the "CSI" header) so they survive the
    # shift. Bid_Form is NOT shifted: its rows 1–3 are title/details/blank, which
    # the RED banner simply overrides.
    if "Leveled_Normalized" in wb.sheetnames:
        ws_lev = wb["Leveled_Normalized"]
        ws_lev.insert_rows(1, 3)
        _write_quarantine_banner(ws_lev, num_contractors, n, start_row=1)
        _mark_failing_cells(ws_lev, "Leveled_Normalized", ordered_leveled, failures)

    # --- AUDIT: append the RED tie-out failure rows + the QUARANTINE summary line ---
    _append_audit_failure_rows(wb, failures)
    _append_audit_quarantine_line(wb, n)

    wb.save(output_path)
    return -1 if n is None else n


# ---------------------------------------------------------------------------
# Primary entry point
# ---------------------------------------------------------------------------

def write_matrix(
    bids: list[NormalizedBid],
    output_path: str | Path,
    run: RunInputs,
    audit_items: Optional[list[AuditItem]] = None,
    leveled_bids: Optional[list[NormalizedBid]] = None,
) -> list[dict]:
    """
    Write normalized bid data into a fresh openpyxl Workbook.

    Project identity (title/address/SF basis) is supplied per-run via ``run``
    (RunInputs) — never hardcoded (M1/M2). ``run.gross_sf`` is the confirmed
    $/SF denominator; ``run.sf_basis_label`` labels the $/SF header.

    Option C writes TWO data sheets:
      * ``Bid_Form`` — the faithful mirror (as-submitted ``bids``), with the
        Col C Normalization Note for each known-firm reclass recommendation.
      * ``Leveled_Normalized`` — the moved-dollar view (``leveled_bids``), with
        the estimator-normalized banner; cross-bid audit signals apply here only.
    Plus a single ``AUDIT`` sheet whose ``View`` column segments the slices.
    When ``leveled_bids`` is None the leveled sheet mirrors ``bids`` (no reclass).

    Returns a list of per-bid summary dicts for reporting (mirror values).
    """
    output_path = Path(output_path)
    gsf = int(run.gross_sf)

    # Step 1: Sort bids (leveled-total ascending). The leveled bids are sorted by
    # the SAME order as the mirror so the two sheets line up cell-for-cell.
    ordered_bids = _sort_bids(bids)
    print(f"  [write_matrix] Contractor order: "
          f"{[b.contractor_name for b in ordered_bids]}")

    leveled_by_name: dict[str, NormalizedBid] = {}
    if leveled_bids is not None:
        leveled_by_name = {b.contractor_name: b for b in leveled_bids}
    ordered_leveled = [
        leveled_by_name.get(b.contractor_name, b) for b in ordered_bids
    ]

    # Step 2: Create fresh workbook + the Bid_Form mirror sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bid_Form"

    # Mirror: as-submitted, Normalization Notes shown, mirror audit slice.
    footer_summaries = _populate_data_sheet(
        ws, ordered_bids, gsf, run, audit_items, view="mirror", show_notes=True
    )

    # Leveled_Normalized: moved dollars applied, banner, leveled audit slice (§3).
    ws_lev = wb.create_sheet(title="Leveled_Normalized")
    _populate_data_sheet(
        ws_lev, ordered_leveled, gsf, run, audit_items,
        view="leveled", show_notes=False,
    )
    _write_leveled_banner(ws_lev, len(ordered_leveled))

    # Step 8: Write AUDIT sheet (single sheet, View column segments the slices)
    if audit_items:
        _write_audit_sheet(wb, audit_items)

    # Step 9: Save
    wb.save(output_path)
    print(f"  [write_matrix] Saved fresh workbook → {output_path}")
    print(f"  [write_matrix] Sheet dimensions: "
          f"{ws.max_row} rows × {ws.max_column} cols")

    # Build per-bid summary dicts (same shape as original for pipeline.py).
    # Row numbers are no longer sequential (dynamic per-item layout), so we
    # report 0 as the row sentinel and let pipeline.py print amounts only.
    summaries: list[dict] = []
    for i, bid in enumerate(ordered_bids):
        # Aggregate subtotal per CSI code (handles duplicate codes)
        seen: dict[str, float] = {}
        for div in bid.divisions:
            amount = _cell_amount(div.subtotal_cell.state, div.subtotal_cell.amount)
            seen[div.csi_code] = seen.get(div.csi_code, 0.0) + amount

        divisions_written = [
            {
                "csi_code": csi_code,
                "row": 0,  # dynamic layout — row number not fixed
                "amount": seen.get(csi_code, 0.0),
                "state": "AMOUNT",
            }
            for csi_code, _ in DIVISION_ROWS
        ]

        summaries.append({
            "contractor": bid.contractor_name,
            "matched": True,
            "name_col": _col_start(i),
            "divisions_written": divisions_written,
            "footer_written": footer_summaries[i],
            "warnings": [],
        })

    return summaries
